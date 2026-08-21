from pathlib import Path
from openpyxl import load_workbook

source = Path("outputs/01a018f8-24d8-7651-b292-e4dc705bf026/19条产业链岗位与职业匹配表.xlsx")
wb = load_workbook(source, read_only=True, data_only=True)
ws = wb["岗位-职业匹配表"]
job_words = ("电气", "电力", "电机", "电控", "电源", "光伏", "控制", "自动化", "设备", "机电", "传感", "变压器", "电路", "运维", "维修", "调试", "系统集成")
chain_words = ("新能源与电力装备产业链", "高端装备与智能制造产业链", "基础设施与城市建设产业链", "汽车与智能网联汽车产业链", "智能物联与消费电子产业链", "机器人产业链")
current = ["", "", ""]
for idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
    vals = ["" if v is None else str(v) for v in row]
    for pos in range(3):
        if vals[pos]:
            current[pos] = vals[pos]
    job, job_id, chain = current
    occupation, code = vals[3], vals[4]
    if any(word in job for word in job_words) and any(word in chain for word in chain_words):
        print(idx, "||", job, "||", job_id, "||", chain, "||", occupation, "||", code)
wb.close()
