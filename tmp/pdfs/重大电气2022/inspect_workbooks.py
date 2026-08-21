from pathlib import Path
import json
from openpyxl import load_workbook

root = Path("tmp/pdfs/重大电气2022")
sources = [
    Path("outputs/01a018f8-24d8-7651-b292-e4dc705bf026/19条产业链岗位与职业匹配表.xlsx"),
    Path("outputs/019f7e5b-0d35-7c91-a1ff-e290f4069b8a/岗位与产业节点关联表.xlsx"),
    Path("岗位详情字段爬取模板.xlsx"),
]

keywords = (
    "电气", "电力", "电机", "电源", "变电", "输电", "配电", "继电", "能源", "充电",
    "自动化", "控制", "运维", "设备", "测试", "调试", "故障", "建筑智能", "传感", "电网",
)
target_chains = (
    "新能源与电力装备产业链", "高端装备与智能制造产业链", "基础设施与城市建设产业链",
    "汽车与智能网联汽车产业链", "智能物联与消费电子产业链", "机器人产业链",
)

report = []
for source in sources:
    wb = load_workbook(source, read_only=True, data_only=False)
    entry = {"path": str(source), "sheets": []}
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, ())
        header_text = ["" if value is None else str(value) for value in header]
        selected = []
        for idx, row in enumerate(rows, start=2):
            values = ["" if value is None else str(value) for value in row]
            joined = " | ".join(values)
            if any(chain in joined for chain in target_chains) or any(word in joined for word in keywords):
                selected.append({"row": idx, "values": values})
        entry["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "header": header_text,
            "selected_count": len(selected),
            "selected": selected,
        })
    wb.close()
    report.append(entry)

output = root / "workbook_inspection.json"
output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
print(output)
for entry in report:
    print("\n", entry["path"])
    for sheet in entry["sheets"]:
        print(sheet["title"], sheet["max_row"], sheet["max_column"], sheet["selected_count"])
        print(sheet["header"][:16])
