from pathlib import Path
from openpyxl import load_workbook

source = Path("outputs/01a018f8-24d8-7651-b292-e4dc705bf026/19条产业链岗位与职业匹配表.xlsx")
wb = load_workbook(source, read_only=True, data_only=True)
for ws in wb.worksheets:
    print(f"\n### {ws.title}")
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = ["" if v is None else str(v) for v in row]
        if idx <= 8:
            print(idx, " || ".join(values))
        if ws.title == "岗位-职业匹配表" and "新能源与电力装备产业链" in values:
            print("TARGET", idx, " || ".join(values))
        if ws.title == "岗位-职业匹配表" and 335 <= idx <= 365:
            print("CHAIN_WINDOW", idx, " || ".join(values))
        if ws.title == "岗位-职业匹配表" and any(k in "|".join(values) for k in ("电气工程师", "电力工程师", "新能源电机工程师", "自动控制工程师")):
            print("ELECTRIC", idx, " || ".join(values))
wb.close()
