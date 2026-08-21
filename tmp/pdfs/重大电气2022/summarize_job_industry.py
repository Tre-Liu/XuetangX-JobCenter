from pathlib import Path
from openpyxl import load_workbook

source = Path("outputs/019f7e5b-0d35-7c91-a1ff-e290f4069b8a/岗位与产业节点关联表.xlsx")
wb = load_workbook(source, read_only=True, data_only=True)
terms = ("电气工程师", "电力工程师", "新能源电机工程师", "新能源电控工程师", "电池/电源开发", "光伏系统工程师", "自动控制工程师", "电气/电器工程师", "变压器与磁电工程师")
for ws in wb.worksheets:
    print(f"\n### {ws.title}")
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = ["" if v is None else str(v) for v in row]
        joined = " || ".join(values)
        if idx <= 8:
            print(idx, joined)
        if any(term in joined for term in terms):
            print("TARGET", idx, joined)
wb.close()
