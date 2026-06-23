from __future__ import annotations

import html
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl


SOURCE_XLSX = Path("V1.0需求（2026.6.11）/官方数据/DW_专业建设数据模型设计_岗位所属行业更新.xlsx")
OUT_DIR = Path("output/dw-er")
OUT_HTML = OUT_DIR / "dw_er.html"
OUT_JSON = OUT_DIR / "dw_schema.json"
OUT_COVERAGE = OUT_DIR / "coverage.md"


SHEET_TO_OBJECT = {
    "01_专业库": ("major", "专业", "专业库", "专业主数据，贯穿初始化、课程、岗位和报告"),
    "02_行业库": ("industry", "行业", "行业库", "国民经济行业分类及行业描述"),
    "03_产业链库": ("industry_chain", "产业链", "产业链库", "专业建设匹配的产业链主数据"),
    "04_产业环节库": ("industry_node", "产业环节", "产业环节库", "产业链下的环节、节点和场景"),
    "04A_产业环节关系库": ("industry_node_relation", "产业环节关系", "产业环节关系库", "产业环节之间的上下游、支撑、协同关系和桑基图边"),
    "05_职业库": ("occupation", "职业", "职业库", "国家职业分类大典职业主数据"),
    "06_岗位库": ("job", "岗位", "岗位库", "平台岗位画像和岗位归一主实体"),
    "06A_岗位资源匹配库": ("job_resource_match", "岗位资源匹配", "岗位资源匹配库", "岗位画像中专业、证书、企业等资源匹配结果"),
    "07_岗位任务库": ("job_task", "岗位任务", "岗位任务库", "岗位典型工作任务明细"),
    "08_岗位能力库": ("job_ability", "岗位能力", "岗位能力库", "岗位知识、技能、素养能力项"),
    "09_企业库": ("company", "企业", "企业库", "企业主数据和工商/行业归一信息"),
    "10_招聘信息库": ("job_posting", "招聘信息", "招聘信息库", "原始招聘信息和岗位/企业归一结果"),
    "11_招聘趋势库": ("recruitment_trend", "招聘趋势", "招聘趋势库", "按周期沉淀的招聘趋势统计"),
    "11A_新技术岗位匹配库": ("emerging_technology_match", "新技术岗位匹配", "新技术岗位匹配库", "新技术方向、岗位、专业、课程和能力建议的复核结果"),
    "12_赛事库": ("competition", "赛事", "赛事库", "专业、岗位能力相关赛事资源"),
    "13_证书库": ("certificate", "证书", "证书库", "职业资格、技能等级和 1+X 证书"),
    "14_课程库": ("course", "课程", "课程库", "课程主数据及专业、能力映射基础"),
    "15_政策库": ("policy", "政策", "政策库", "政策文件、新闻与建设依据"),
    "16_初始化批次库": ("init_batch", "初始化批次", "初始化批次库", "专业初始化任务、默认链路和状态"),
    "17_数据来源目录": ("source_catalog", "数据来源", "数据来源目录", "DW 模型使用的数据来源目录"),
}


POSITIONS = {
    "source_catalog": (60, 40),
    "major": (60, 290),
    "init_batch": (60, 540),
    "policy": (60, 790),
    "industry_chain": (470, 170),
    "industry_node": (470, 500),
    "industry": (470, 830),
    "industry_node_relation": (880, 590),
    "job": (880, 330),
    "occupation": (880, 80),
    "job_resource_match": (1290, 20),
    "job_task": (1290, 300),
    "job_ability": (1290, 580),
    "course": (1290, 860),
    "certificate": (1290, 1140),
    "company": (1700, 230),
    "job_posting": (1700, 560),
    "recruitment_trend": (1700, 890),
    "emerging_technology_match": (2110, 690),
    "competition": (60, 1040),
}


FIELD_OVERRIDES = {
    ("rel_chain_node", "industry_node"): (["chain_id"], ["chain_id"]),
    ("rel_chain_node_relation", "industry_node_relation"): (["chain_id"], ["chain_id"]),
    ("rel_node_relation_source", "industry_node_relation"): (["industry_node_id"], ["source_node_id"]),
    ("rel_node_relation_target", "industry_node"): (["target_node_id"], ["industry_node_id"]),
    ("rel_node_job", "job"): (["industry_node_id"], ["industry_node_id"]),
    ("rel_job_task", "job_task"): (["job_id"], ["job_id"]),
    ("rel_job_ability", "job_ability"): (["job_id"], ["job_id"]),
    ("rel_task_ability", "job_ability"): (["ability_ids"], ["ability_id"]),
    ("rel_posting_job", "job"): (["normalized_job_id"], ["job_id"]),
    ("rel_posting_company", "company"): (["company_id"], ["company_id"]),
    ("rel_trend_job_job", "job"): (["job_id"], ["job_id"]),
    ("rel_trend_job_industry_chain", "industry_chain"): (["chain_id"], ["chain_id"]),
    ("rel_competition_major", "major"): (["major_codes"], ["major_code"]),
    ("rel_competition_ability", "job_ability"): (["ability_requirements"], ["ability_id"]),
    ("rel_policy_context_industry_chain", "industry_chain"): (["topic_tags"], ["chain_id"]),
    ("rel_policy_context_major", "major"): (["topic_tags"], ["major_id"]),
    ("rel_policy_context_job", "job"): (["topic_tags"], ["job_id"]),
}


ADDITIONAL_RELATIONS = [
    ("rel_init_major", "初始化批次", "专业", "N:1", "major_id", "初始化批次关联被初始化的专业。"),
    ("rel_init_chain", "初始化批次", "产业链", "N:1", "selected_chain_id", "初始化批次记录管理员选择的默认产业链。"),
]


@dataclass
class Field:
    code: str
    name: str
    data_type: str
    required: str
    key: str
    description: str
    example: str
    source: str
    frequency: str
    note: str


@dataclass
class DbObject:
    key: str
    label: str
    sheet: str
    title: str
    meaning: str
    fields: list[Field]


@dataclass
class Relation:
    code: str
    source_key: str
    source_label: str
    source_fields: list[str]
    target_key: str
    target_label: str
    target_fields: list[str]
    relation_type: str
    description: str


def esc(value: object) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def read_objects() -> tuple[list[DbObject], list[dict[str, str]]]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    objects: list[DbObject] = []
    sources: list[dict[str, str]] = []
    for sheet_name, (key, label, title, meaning) in SHEET_TO_OBJECT.items():
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if sheet_name == "17_数据来源目录":
            headers = [cell(v) for v in rows[0]]
            for row in rows[1:]:
                item = dict(zip(headers, [cell(v) for v in row]))
                if any(item.values()):
                    sources.append(item)
            fields = [
                Field("module", "模块", "string", "是", "IDX", "来源所属模块", "", "", "按来源维护", ""),
                Field("source_name", "来源名称", "string", "是", "IDX", "官方、行业或学校来源名称", "", "", "按来源维护", ""),
                Field("url", "网址", "url", "否", "", "来源访问地址", "", "", "按来源维护", ""),
                Field("usage", "用途", "text", "否", "", "该来源支撑的数据口径", "", "", "按来源维护", ""),
            ]
        else:
            fields = []
            for row in rows[1:]:
                if not row or not cell(row[0]):
                    continue
                values = [cell(v) for v in row[:10]]
                values.extend([""] * (10 - len(values)))
                fields.append(Field(*values[:10]))
        objects.append(DbObject(key, label, sheet_name, title, meaning, fields))
    return objects, sources


def object_maps(objects: list[DbObject]) -> tuple[dict[str, DbObject], dict[str, str]]:
    by_key = {obj.key: obj for obj in objects}
    by_label = {obj.label: obj.key for obj in objects}
    return by_key, by_label


def pk_fields(obj: DbObject) -> list[str]:
    fields = [field.code for field in obj.fields if "PK" in field.key.upper()]
    return fields or ([obj.fields[0].code] if obj.fields else [])


def parse_relation_fields(text: str, source_obj: DbObject, target_obj: DbObject) -> tuple[list[str], list[str]]:
    text = cell(text)
    if "->" in text:
        left, right = [part.strip() for part in text.split("->", 1)]
        source_fields = [part.strip() for part in re.split(r"[,，/]", left) if part.strip()]
        target_fields = [part.strip() for part in re.split(r"[,，/]", right) if part.strip()]
        if len(target_fields) == 1 and target_fields[0] == target_obj.key:
            target_fields = pk_fields(target_obj)
        return source_fields, target_fields

    parts = [part.strip() for part in re.split(r"[,，+]", text) if part.strip()]
    source_candidates = {field.code for field in source_obj.fields}
    target_candidates = {field.code for field in target_obj.fields}
    source_fields = [part for part in parts if part in source_candidates]
    target_fields = [part for part in parts if part in target_candidates and part not in source_fields]
    if not source_fields:
        source_fields = [part for part in parts if part.endswith("_id") or part.endswith("_code")][:1]
    if not target_fields:
        target_fields = pk_fields(target_obj)
    return source_fields, target_fields


def read_relations(objects: list[DbObject]) -> list[Relation]:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["18_关联关系"]
    by_key, by_label = object_maps(objects)
    relations: list[Relation] = []
    rows = list(ws.iter_rows(min_row=2, values_only=True)) + ADDITIONAL_RELATIONS
    for row in rows:
        code, source_label, target_label_text, rel_type, rel_fields, description = [cell(v) for v in row[:6]]
        if not code or source_label not in by_label:
            continue
        source_key = by_label[source_label]
        target_labels = [part.strip() for part in re.split(r"[/、]", target_label_text) if part.strip()]
        for target_label in target_labels:
            if target_label not in by_label:
                continue
            target_key = by_label[target_label]
            source_fields, target_fields = parse_relation_fields(rel_fields, by_key[source_key], by_key[target_key])
            source_fields, target_fields = FIELD_OVERRIDES.get((code if len(target_labels) == 1 else f"{code}_{target_key}", target_key), (source_fields, target_fields))
            relations.append(
                Relation(
                    code=code if len(target_labels) == 1 else f"{code}_{target_key}",
                    source_key=source_key,
                    source_label=source_label,
                    source_fields=source_fields,
                    target_key=target_key,
                    target_label=target_label,
                    target_fields=target_fields,
                    relation_type=rel_type,
                    description=description,
                )
            )
    return relations


def key_fields(obj: DbObject, relations: list[Relation], limit: int = 8) -> tuple[list[Field], int]:
    relation_fields = set()
    for rel in relations:
        if rel.source_key == obj.key:
            relation_fields.update(rel.source_fields)
        if rel.target_key == obj.key:
            relation_fields.update(rel.target_fields)
    ordered: list[Field] = []

    def add(predicate) -> None:
        for field in obj.fields:
            if predicate(field) and field.code not in {item.code for item in ordered}:
                ordered.append(field)

    add(lambda field: "PK" in field.key.upper())
    add(lambda field: field.code in relation_fields)
    add(lambda field: "FK" in field.key.upper())
    add(lambda field: "IDX" in field.key.upper() or "UK" in field.key.upper())
    add(lambda _field: True)
    visible = ordered[:limit]
    return visible, max(0, len(obj.fields) - len(visible))


def field_tag(field: Field, obj: DbObject, relations: list[Relation]) -> str:
    if "PK" in field.key.upper():
        return "PK"
    for rel in relations:
        if rel.source_key == obj.key and field.code in rel.source_fields:
            return "FK"
        if rel.target_key == obj.key and field.code in rel.target_fields:
            return "REF"
    if "FK" in field.key.upper():
        return "FK"
    if "UK" in field.key.upper():
        return "UK"
    if "IDX" in field.key.upper():
        return "IDX"
    return ""


def trim(text: str, limit: int) -> str:
    return text if len(text) <= limit else text[: limit - 1] + "…"


def render_svg(objects: list[DbObject], relations: list[Relation]) -> str:
    w, h = 2100, 1230
    node_w, node_h = 320, 185
    object_keys = {obj.key for obj in objects}
    parts = [
        f'<svg class="graph" viewBox="0 0 {w} {h}" role="img" aria-label="DW relationship graph">',
        '<defs><marker id="arrow-solid" markerWidth="8" markerHeight="8" refX="7" refY="4" orient="auto"><path d="M0,0 L8,4 L0,8 z" fill="#2563eb"/></marker><marker id="arrow-dashed" markerWidth="8" markerHeight="8" refX="7" refY="4" orient="auto"><path d="M0,0 L8,4 L0,8 z" fill="#64748b"/></marker></defs>',
    ]
    for idx, rel in enumerate(relations, 1):
        if rel.source_key not in POSITIONS or rel.target_key not in POSITIONS:
            continue
        sx, sy = POSITIONS[rel.source_key]
        tx, ty = POSITIONS[rel.target_key]
        x1, y1 = sx + node_w, sy + node_h / 2
        x2, y2 = tx, ty + node_h / 2
        if tx < sx:
            x1, x2 = sx, tx + node_w
        dx = max(110, abs(x2 - x1) / 2)
        solid = "1:" in rel.relation_type or "N:1" in rel.relation_type or "1:N" in rel.relation_type
        cls = "edge explicit" if solid else "edge inferred"
        marker = "arrow-solid" if solid else "arrow-dashed"
        title = f"{rel.source_label}.{', '.join(rel.source_fields)} -> {rel.target_label}.{', '.join(rel.target_fields)}"
        parts.append(
            f'<path class="{cls}" data-rel="r{idx}" data-source="{esc(rel.source_key)}" data-target="{esc(rel.target_key)}" '
            f'd="M{x1},{y1} C{x1+dx},{y1} {x2-dx},{y2} {x2},{y2}" marker-end="url(#{marker})">'
            f"<title>{esc(title)}</title></path>"
        )
    for obj in objects:
        if obj.key not in POSITIONS:
            continue
        x, y = POSITIONS[obj.key]
        fields, remaining = key_fields(obj, relations)
        klass = "node reference" if obj.key == "source_catalog" else "node table"
        parts.append(f'<g class="{klass}" data-name="{esc(obj.key)}">')
        parts.append(f'<rect x="{x}" y="{y}" width="{node_w}" height="{node_h}" rx="8"/>')
        parts.append(f'<text class="node-title" x="{x+14}" y="{y+25}">{esc(obj.title)}</text>')
        parts.append(f'<text class="node-key" x="{x+14}" y="{y+45}">{esc(obj.key)}</text>')
        parts.append(f'<text class="node-meaning" x="{x+14}" y="{y+65}">{esc(trim(obj.meaning, 34))}</text>')
        parts.append(f'<line class="node-rule" x1="{x+14}" y1="{y+78}" x2="{x+node_w-14}" y2="{y+78}"/>')
        fy = y + 98
        for field in fields:
            tag = field_tag(field, obj, relations)
            prefix = f"[{tag}] " if tag else ""
            label = f"{field.code}: {field.data_type}"
            parts.append(f'<text class="node-field" x="{x+14}" y="{fy}">{esc(prefix + trim(label, 38))}</text>')
            fy += 14
        if remaining:
            parts.append(f'<text class="node-more" x="{x+14}" y="{fy}">+ {remaining} fields below</text>')
        parts.append("</g>")
    parts.append("</svg>")
    return "\n".join(parts)


def render_relation_rows(relations: list[Relation]) -> str:
    rows = []
    for idx, rel in enumerate(relations, 1):
        rows.append(
            f'<tr data-rel="r{idx}" data-source="{esc(rel.source_key)}" data-target="{esc(rel.target_key)}">'
            f"<td>{idx}</td><td><code>{esc(rel.source_label)}.{esc(', '.join(rel.source_fields))}</code></td>"
            f"<td>-></td><td><code>{esc(rel.target_label)}.{esc(', '.join(rel.target_fields))}</code></td>"
            f"<td><span class=\"badge inferred\">{esc(rel.relation_type)}</span></td><td>{esc(rel.description)}</td></tr>"
        )
    return "\n".join(rows)


def render_cards(objects: list[DbObject], relations: list[Relation]) -> str:
    cards = []
    for obj in objects:
        rows = []
        for field in obj.fields:
            tag = field_tag(field, obj, relations) or field.key
            tags = " ".join(f'<span class="mini">{esc(part)}</span>' for part in re.split(r"[/,， ]+", tag) if part)
            rows.append(
                f"<tr><td><code>{esc(field.code)}</code></td><td>{esc(field.name)}</td><td>{esc(field.data_type)}</td>"
                f"<td>{tags}</td><td>{esc(field.description)}</td><td>{esc(field.note)}</td></tr>"
            )
        cards.append(
            f'<section class="card" data-name="{esc(obj.key)}"><h3>{esc(obj.title)} <span>{esc(obj.key)}</span></h3>'
            f"<p>{esc(obj.meaning)}</p><table class=\"fields\"><thead><tr><th>字段编码</th><th>中文名称</th><th>类型</th><th>键/索引</th><th>说明</th><th>关联备注</th></tr></thead>"
            f"<tbody>{''.join(rows)}</tbody></table></section>"
        )
    return "\n".join(cards)


def render_hover_css(objects: list[DbObject], relations: list[Relation]) -> str:
    rules = [
        ".graph:has(.node:hover) .edge, .graph:has(.edge:hover) .edge { opacity: .06; }",
        ".graph:has(.node:hover) .node, .graph:has(.edge:hover) .node { opacity: .25; }",
        ".graph .node:hover { opacity: 1; }",
        ".graph .node:hover rect { stroke: #2563eb; stroke-width: 2.6; fill: #eff6ff; }",
        ".graph .edge:hover { opacity: 1; stroke-width: 4.5; }",
    ]
    for obj in objects:
        edge_selectors = [
            f'.graph:has(.node[data-name="{obj.key}"]:hover) .edge[data-source="{obj.key}"]',
            f'.graph:has(.node[data-name="{obj.key}"]:hover) .edge[data-target="{obj.key}"]',
        ]
        rules.append(", ".join(edge_selectors) + " { opacity: 1; stroke-width: 4.5; }")
        connected = {obj.key}
        for rel in relations:
            if rel.source_key == obj.key:
                connected.add(rel.target_key)
            if rel.target_key == obj.key:
                connected.add(rel.source_key)
        node_selectors = [f'.graph:has(.node[data-name="{obj.key}"]:hover) .node[data-name="{key}"]' for key in sorted(connected)]
        rules.append(", ".join(node_selectors) + " { opacity: 1; }")
        rules.append(", ".join(f"{selector} rect" for selector in node_selectors) + " { stroke: #2563eb; stroke-width: 2.6; }")
    for idx, rel in enumerate(relations, 1):
        rules.append(f'.graph:has(.edge[data-rel="r{idx}"]:hover) .edge[data-rel="r{idx}"] {{ opacity: 1; stroke-width: 4.5; }}')
        rules.append(
            f'.graph:has(.edge[data-rel="r{idx}"]:hover) .node[data-name="{rel.source_key}"], '
            f'.graph:has(.edge[data-rel="r{idx}"]:hover) .node[data-name="{rel.target_key}"] {{ opacity: 1; }}'
        )
        rules.append(
            f'.graph:has(.edge[data-rel="r{idx}"]:hover) .node[data-name="{rel.source_key}"] rect, '
            f'.graph:has(.edge[data-rel="r{idx}"]:hover) .node[data-name="{rel.target_key}"] rect {{ stroke: #2563eb; stroke-width: 2.6; }}'
        )
    return "\n    ".join(rules)


def render_html(objects: list[DbObject], relations: list[Relation], sources: list[dict[str, str]]) -> str:
    table_count = len(objects)
    source_count = len(sources)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>DW 专业建设数据模型关系图</title>
  <style>
    :root {{ --bg:#f7f8fb; --ink:#1f2937; --muted:#64748b; --line:#d7dde8; --blue:#2563eb; }}
    * {{ box-sizing: border-box; }}
    body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif; color:var(--ink); background:var(--bg); }}
    header {{ padding:28px 34px 18px; background:#111827; color:white; }}
    h1 {{ margin:0 0 10px; font-size:28px; letter-spacing:0; }}
    header p {{ margin:5px 0; color:#cbd5e1; line-height:1.55; }}
    main {{ padding:22px 34px 50px; }}
    input {{ width:min(560px,100%); padding:11px 13px; border:1px solid var(--line); border-radius:8px; font-size:14px; background:white; }}
    .stats {{ display:grid; grid-template-columns:repeat(4,minmax(150px,1fr)); gap:12px; margin:16px 0 22px; }}
    .stat {{ background:white; border:1px solid var(--line); border-radius:8px; padding:14px 16px; }}
    .stat strong {{ display:block; font-size:24px; margin-bottom:2px; }}
    .stat span {{ color:var(--muted); font-size:13px; }}
    .panel {{ background:white; border:1px solid var(--line); border-radius:8px; padding:16px; margin:18px 0; overflow:auto; }}
    h2 {{ font-size:19px; margin:0 0 13px; }}
    .legend {{ display:flex; gap:18px; flex-wrap:wrap; color:var(--muted); font-size:13px; margin-bottom:10px; }}
    .legend i {{ display:inline-block; width:34px; height:0; border-top:3px solid var(--blue); vertical-align:middle; margin-right:7px; }}
    .legend .dash {{ border-top-style:dashed; border-color:#64748b; }}
    .graph {{ min-width:1600px; width:100%; height:auto; background:#fbfdff; border:1px solid #e5eaf2; border-radius:8px; }}
    .edge {{ fill:none; stroke-width:1.7; opacity:.35; pointer-events:stroke; transition:opacity .12s ease, stroke-width .12s ease; }}
    .edge.explicit {{ stroke:var(--blue); opacity:.66; stroke-width:2.6; }}
    .edge.inferred {{ stroke:#64748b; stroke-dasharray:8 7; }}
    .edge.is-dimmed {{ opacity:.06; }}
    .edge.is-highlighted {{ opacity:1; stroke-width:4.5; }}
    .node {{ transition:opacity .12s ease; }}
    .node rect {{ fill:white; stroke:#cbd5e1; stroke-width:1.2; filter:drop-shadow(0 1px 2px rgba(15,23,42,.08)); transition:stroke .12s ease, stroke-width .12s ease, fill .12s ease; }}
    .node.reference rect {{ fill:#f8fafc; }}
    .node.is-dimmed {{ opacity:.25; }}
    .node.is-highlighted rect {{ stroke:var(--blue); stroke-width:2.6; fill:#eff6ff; }}
    .node-title {{ font-size:15px; font-weight:700; fill:#111827; }}
    .node-key,.node-meaning,.node-more {{ font-size:11px; fill:#64748b; }}
    .node-rule {{ stroke:#e2e8f0; stroke-width:1; }}
    .node-field {{ font-size:11px; fill:#0f172a; font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; }}
    table {{ width:100%; border-collapse:collapse; }}
    th,td {{ padding:9px 10px; border-bottom:1px solid #e7ecf4; text-align:left; vertical-align:top; font-size:13px; line-height:1.45; }}
    th {{ color:#475569; background:#f8fafc; position:sticky; top:0; }}
    code {{ font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; font-size:12px; color:#0f172a; }}
    .badge,.mini {{ display:inline-flex; align-items:center; border-radius:999px; padding:2px 8px; font-size:12px; white-space:nowrap; background:#e2e8f0; color:#475569; margin-right:4px; }}
    .cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(520px,1fr)); gap:14px; }}
    .card {{ background:white; border:1px solid var(--line); border-radius:8px; padding:14px; overflow:hidden; }}
    .card h3 {{ margin:0 0 5px; font-size:16px; }}
    .card h3 span {{ color:var(--muted); font-size:12px; font-weight:500; margin-left:8px; }}
    .card p {{ color:var(--muted); margin:0 0 10px; font-size:13px; line-height:1.5; }}
    #relationTable tbody tr {{ transition:background .12s ease, opacity .12s ease; }}
    #relationTable tbody tr.is-highlighted {{ background:#eff6ff; }}
    #relationTable tbody tr.is-dimmed {{ opacity:.38; }}
    {render_hover_css(objects, relations)}
  </style>
</head>
<body>
  <header>
    <h1>DW 专业建设数据模型关系图</h1>
    <p>来源 Excel: {esc(SOURCE_XLSX)}。依据字段字典与 18_关联关系重绘；节点间距已加大，线条支持鼠标 hover 高亮。</p>
    <p>主链路: 专业 -> 产业链 -> 产业环节/产业环节关系 -> 岗位 -> 任务/能力/课程/证书；桑基边由产业环节关系库承载。</p>
  </header>
  <main>
    <input id="filter" type="search" placeholder="搜索表、字段或关系，例如 岗位 / company_id / ability_id">
    <section class="stats">
      <div class="stat"><strong>{table_count}</strong><span>DW 对象节点</span></div>
      <div class="stat"><strong>{len(relations)}</strong><span>字段级关系</span></div>
      <div class="stat"><strong>{source_count}</strong><span>数据来源条目</span></div>
      <div class="stat"><strong>{sum(len(obj.fields) for obj in objects)}</strong><span>字段总数</span></div>
    </section>
    <section class="panel">
      <h2>关系总览图</h2>
      <div class="legend"><span><i></i>主从/归一关系</span><span><i class="dash"></i>N:M/聚合/上下文关系</span></div>
      {render_svg(objects, relations)}
    </section>
    <section class="panel">
      <h2>字段级关系列表</h2>
      <table id="relationTable"><thead><tr><th>#</th><th>来源字段</th><th></th><th>目标字段</th><th>类型</th><th>业务说明</th></tr></thead><tbody>{render_relation_rows(relations)}</tbody></table>
    </section>
    <section>
      <h2>DW 表字段明细</h2>
      <div class="cards" id="cards">{render_cards(objects, relations)}</div>
    </section>
  </main>
  <script>
    const filter = document.getElementById('filter');
    const cards = Array.from(document.querySelectorAll('.card'));
    const graphNodes = Array.from(document.querySelectorAll('g.node'));
    const edges = Array.from(document.querySelectorAll('.edge'));
    const relRows = Array.from(document.querySelectorAll('#relationTable tbody tr'));
    function setGraphState(activeNode, activeRel) {{
      const connectedNodes = new Set();
      if (activeNode) connectedNodes.add(activeNode);
      edges.forEach(edge => {{
        const source = edge.dataset.source, target = edge.dataset.target, rel = edge.dataset.rel;
        const on = (activeNode && (source === activeNode || target === activeNode)) || (activeRel && rel === activeRel);
        edge.classList.toggle('is-highlighted', on);
        edge.classList.toggle('is-dimmed', Boolean(activeNode || activeRel) && !on);
        if (on) {{ connectedNodes.add(source); connectedNodes.add(target); }}
      }});
      graphNodes.forEach(node => {{
        const on = connectedNodes.has(node.dataset.name);
        node.classList.toggle('is-highlighted', on);
        node.classList.toggle('is-dimmed', Boolean(activeNode || activeRel) && !on);
      }});
      relRows.forEach(row => {{
        const on = (activeRel && row.dataset.rel === activeRel) || (activeNode && (row.dataset.source === activeNode || row.dataset.target === activeNode));
        row.classList.toggle('is-highlighted', on);
        row.classList.toggle('is-dimmed', Boolean(activeNode || activeRel) && !on);
      }});
    }}
    function clearGraphState() {{ setGraphState('', ''); }}
    function bindHover(el, enter, leave) {{
      ['mouseenter','mouseover','pointerenter'].forEach(type => el.addEventListener(type, enter));
      ['mouseleave','mouseout','pointerleave'].forEach(type => el.addEventListener(type, leave));
    }}
    graphNodes.forEach(node => bindHover(node, () => setGraphState(node.dataset.name, ''), clearGraphState));
    edges.forEach(edge => bindHover(edge, () => setGraphState('', edge.dataset.rel), clearGraphState));
    relRows.forEach(row => bindHover(row, () => setGraphState('', row.dataset.rel), clearGraphState));
    cards.forEach(card => bindHover(card, () => setGraphState(card.dataset.name, ''), clearGraphState));
    filter.addEventListener('input', () => {{
      const q = filter.value.trim().toLowerCase();
      cards.forEach(card => card.style.display = card.innerText.toLowerCase().includes(q) ? '' : 'none');
      relRows.forEach(row => row.style.display = row.innerText.toLowerCase().includes(q) ? '' : 'none');
      clearGraphState();
    }});
  </script>
</body>
</html>"""


def write_outputs(objects: list[DbObject], relations: list[Relation], sources: list[dict[str, str]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_HTML.write_text(render_html(objects, relations, sources), encoding="utf-8")
    OUT_JSON.write_text(
        json.dumps(
            {
                "source_xlsx": str(SOURCE_XLSX),
                "objects": [asdict(obj) for obj in objects],
                "relations": [asdict(rel) for rel in relations],
                "sources": sources,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    lines = [
        "# DW 专业建设数据模型 ER 图覆盖清单",
        "",
        f"- 来源 Excel: `{SOURCE_XLSX}`",
        f"- DW 对象节点: {len(objects)}",
        f"- 字段总数: {sum(len(obj.fields) for obj in objects)}",
        f"- 字段级关系: {len(relations)}",
        f"- 数据来源条目: {len(sources)}",
        "",
        "## 对象",
    ]
    for obj in objects:
        lines.append(f"- `{obj.key}` ({obj.title}): {len(obj.fields)} 字段")
    lines.extend(["", "## 关系"])
    for rel in relations:
        lines.append(
            f"- `{rel.source_label}.{', '.join(rel.source_fields)}` -> `{rel.target_label}.{', '.join(rel.target_fields)}` ({rel.relation_type})"
        )
    OUT_COVERAGE.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    objects, sources = read_objects()
    relations = read_relations(objects)
    if len(objects) < 16 or not relations:
        raise SystemExit(f"Unexpected DW extraction: objects={len(objects)} relations={len(relations)}")
    write_outputs(objects, relations, sources)
    print(f"Wrote {OUT_HTML}")
    print(f"Wrote {OUT_JSON}")
    print(f"Wrote {OUT_COVERAGE}")


if __name__ == "__main__":
    main()
