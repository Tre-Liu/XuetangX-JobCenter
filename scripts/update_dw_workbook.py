from __future__ import annotations

from copy import copy
from pathlib import Path
import shutil

import openpyxl


DATA_DIR = Path("/Users/liuhongzhe/Desktop/学堂/专业建设/Codex工程/V1.0需求（2026.6.11）/官方数据")
ORIGINAL_XLSX = DATA_DIR / "【DW】专业建设数据模型设计.xlsx"
UPDATED_XLSX = DATA_DIR / "DW_专业建设数据模型设计_岗位所属行业更新.xlsx"
OUT_DIR = Path("output/dw-er")


JOB_FIELDS = [
    [
        "industry_code",
        "所属行业代码",
        "string",
        "是",
        "FK",
        "岗位所属行业代码，按岗位归属行业映射到国民经济行业分类。",
        "I65",
        "行业库/招聘归并/人工维护",
        "按批次",
        "关联02_行业库",
    ],
    [
        "industry_name",
        "所属行业名称",
        "string",
        "否",
        "IDX",
        "岗位所属行业展示名称，冗余自行业库用于页面展示和筛选。",
        "软件和信息技术服务业",
        "行业库/规则映射",
        "按批次",
        "冗余展示字段",
    ],
]

MATCH_SHEET = "04B_专业产业环节匹配库"

MATCH_FIELDS = [
    [
        "match_id",
        "匹配ID",
        "string",
        "是",
        "PK",
        "专业与产业环节匹配记录的唯一标识。",
        "match_maj_node_bim",
        "系统生成",
        "每次初始化/复核",
        "",
    ],
    [
        "major_id",
        "专业ID",
        "string",
        "是",
        "FK",
        "被匹配的专业主键。",
        "maj_460301",
        "专业库/初始化批次",
        "每次初始化/复核",
        "关联01_专业库.major_id",
    ],
    [
        "major_code",
        "专业代码",
        "string",
        "否",
        "IDX",
        "教育部专业目录代码，便于跨系统展示和导入导出。",
        "460301",
        "专业库",
        "随专业维护",
        "冗余展示字段",
    ],
    [
        "industry_node_id",
        "产业环节ID",
        "string",
        "是",
        "FK",
        "专业重点匹配的产业环节或产业节点。",
        "node_bim_platform",
        "产业环节库/初始化算法/人工复核",
        "每次初始化/复核",
        "关联04_产业环节库.industry_node_id",
    ],
    [
        "chain_id",
        "产业链ID",
        "string",
        "是",
        "FK",
        "匹配环节所属产业链，用于按链条过滤专业环节匹配结果。",
        "chain_smart_construction",
        "产业环节库",
        "每次初始化/复核",
        "关联03_产业链库.chain_id",
    ],
    [
        "match_score",
        "环节匹配度",
        "number",
        "否",
        "IDX",
        "专业与产业环节的匹配评分，可综合就业面向、课程、岗位和政策证据。",
        "88",
        "初始化算法/专家复核",
        "每次初始化/复核",
        "",
    ],
    [
        "match_reason",
        "匹配依据说明",
        "text",
        "否",
        "",
        "说明该专业为什么对应此产业环节，支撑评审和人工复核。",
        "核心课程和就业岗位集中承接BIM协同设计与算量平台环节。",
        "AI生成后审核/专家复核",
        "每次初始化/复核",
        "",
    ],
    [
        "evidence_tags",
        "证据标签",
        "text",
        "否",
        "",
        "支撑匹配的课程、岗位、技术、企业或政策关键词。",
        "BIM、工程算量、Revit、模型审查",
        "专业库/岗位库/课程库/政策库",
        "每次初始化/复核",
        "",
    ],
    [
        "is_primary_node",
        "是否重点环节",
        "boolean",
        "否",
        "IDX",
        "标识该产业环节是否为专业初始化或评审中的重点承接环节。",
        "true",
        "初始化算法/人工选择",
        "每次初始化/复核",
        "",
    ],
    [
        "source_batch_id",
        "来源批次ID",
        "string",
        "否",
        "FK",
        "记录该匹配来自哪次初始化、导入或复核批次。",
        "batch_202606",
        "初始化批次库",
        "按批次",
        "关联16_初始化批次库.batch_id",
    ],
    [
        "updated_at",
        "更新时间",
        "datetime",
        "否",
        "",
        "匹配记录最近生成或复核时间。",
        "2026-06-23 14:00",
        "系统生成",
        "每次维护",
        "",
    ],
]

RELATION_ROWS = [
    [
        "rel_job_industry",
        "岗位",
        "行业",
        "N:1",
        "industry_code",
        "岗位按国民经济行业分类归属行业，用于行业维度筛选、统计和趋势分析。",
    ],
    [
        "rel_major_node_match_major",
        "专业",
        "专业产业环节匹配",
        "1:N",
        "major_id -> major_id",
        "专业可匹配多个产业环节，匹配库沉淀匹配分、依据和是否重点环节。",
    ],
    [
        "rel_major_node_match_node",
        "专业产业环节匹配",
        "产业环节",
        "N:1",
        "industry_node_id -> industry_node_id",
        "专业产业环节匹配记录指向具体产业环节，用于按专业查看重点产业节点。",
    ],
    [
        "rel_major_node",
        "专业",
        "产业环节",
        "N:M(经匹配库)",
        "major_id -> industry_node_id",
        "图谱展示用派生关系，实际落库由04B_专业产业环节匹配库承载匹配分、依据和重点环节选择。",
    ],
]


def copy_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def ensure_job_fields(ws) -> None:
    existing = {ws.cell(row=row, column=1).value: row for row in range(2, ws.max_row + 1)}
    if all(field[0] in existing for field in JOB_FIELDS):
        return

    chain_row = existing.get("chain_id")
    insert_at = (chain_row + 1) if chain_row else ws.max_row + 1
    rows_to_add = [field for field in JOB_FIELDS if field[0] not in existing]
    ws.insert_rows(insert_at, amount=len(rows_to_add))

    template_row = max(2, insert_at - 1)
    for offset, values in enumerate(rows_to_add):
        row_idx = insert_at + offset
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            copy_style(ws.cell(row=template_row, column=col_idx), cell)
            cell.value = value


def ensure_match_sheet(wb) -> None:
    if MATCH_SHEET in wb.sheetnames:
        ws = wb[MATCH_SHEET]
    else:
        after_index = wb.sheetnames.index("04A_产业环节关系库") + 1
        ws = wb.create_sheet(MATCH_SHEET, after_index)
        template_ws = wb["04_产业环节库"]
        for col_idx in range(1, template_ws.max_column + 1):
            copy_style(template_ws.cell(row=1, column=col_idx), ws.cell(row=1, column=col_idx))
            ws.cell(row=1, column=col_idx).value = template_ws.cell(row=1, column=col_idx).value
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = template_ws.column_dimensions[
                template_ws.cell(row=1, column=col_idx).column_letter
            ].width

    existing = {ws.cell(row=row, column=1).value: row for row in range(2, ws.max_row + 1)}
    rows_to_add = [field for field in MATCH_FIELDS if field[0] not in existing]
    if not rows_to_add:
        return

    template_row = 2 if ws.max_row < 2 else ws.max_row
    for values in rows_to_add:
        row_idx = ws.max_row + 1
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            copy_style(ws.cell(row=template_row, column=col_idx), cell)
            cell.value = value


def ensure_relations(ws) -> None:
    existing = {ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)}
    for relation_row in RELATION_ROWS:
        if relation_row[0] in existing:
            continue
        row_idx = ws.max_row + 1
        template_row = max(2, row_idx - 1)
        for col_idx, value in enumerate(relation_row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            copy_style(ws.cell(row=template_row, column=col_idx), cell)
            cell.value = value


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    source_xlsx = ORIGINAL_XLSX if ORIGINAL_XLSX.exists() else UPDATED_XLSX
    if source_xlsx != UPDATED_XLSX:
        shutil.copy2(source_xlsx, UPDATED_XLSX)
    wb = openpyxl.load_workbook(UPDATED_XLSX)
    ensure_job_fields(wb["06_岗位库"])
    ensure_match_sheet(wb)
    ensure_relations(wb["18_关联关系"])
    wb.save(UPDATED_XLSX)
    print(f"Wrote {UPDATED_XLSX}")


if __name__ == "__main__":
    main()
