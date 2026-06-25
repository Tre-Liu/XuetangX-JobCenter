from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "V1.0需求（2026.6.11）" / "官方数据" / "DW_专业建设数据模型设计_岗位所属行业更新.xlsx"
GRAPH_JSON = ROOT / "output" / "dw-er" / "dw_schema.json"
GRAPH_HTML = ROOT / "output" / "dw-er" / "dw_er.html"
RELATION_SHEET = "04A_产业环节关系库"
MATCH_SHEET = "04B_专业产业环节匹配库"
EXPECTED_FIELDS = {
    "relation_id",
    "chain_id",
    "source_node_id",
    "target_node_id",
    "relation_type",
    "relation_value",
    "relation_desc",
    "evidence_tags",
    "source_url",
    "updated_at",
}
EXPECTED_MATCH_FIELDS = {
    "match_id",
    "major_id",
    "major_code",
    "industry_node_id",
    "chain_id",
    "match_score",
    "match_reason",
    "evidence_tags",
    "is_primary_node",
    "source_batch_id",
    "updated_at",
}
EXPECTED_RELATIONS = {
    "rel_node_relation_source",
    "rel_node_relation_target",
    "rel_major_node_match_major",
    "rel_major_node_match_node",
    "rel_major_node",
}


def fail(message: str) -> None:
    print(f"FAIL: {message}", file=sys.stderr)
    raise SystemExit(1)


def main() -> None:
    if not WORKBOOK.exists():
        fail(f"missing workbook: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    if RELATION_SHEET not in wb.sheetnames:
        fail(f"missing sheet: {RELATION_SHEET}")
    if MATCH_SHEET not in wb.sheetnames:
        fail(f"missing sheet: {MATCH_SHEET}")

    ws = wb[RELATION_SHEET]
    field_codes = {str(ws.cell(row=row, column=1).value or "").strip() for row in range(2, ws.max_row + 1)}
    missing_fields = sorted(EXPECTED_FIELDS - field_codes)
    if missing_fields:
        fail(f"{RELATION_SHEET} missing fields: {', '.join(missing_fields)}")

    match_ws = wb[MATCH_SHEET]
    match_field_codes = {str(match_ws.cell(row=row, column=1).value or "").strip() for row in range(2, match_ws.max_row + 1)}
    missing_match_fields = sorted(EXPECTED_MATCH_FIELDS - match_field_codes)
    if missing_match_fields:
        fail(f"{MATCH_SHEET} missing fields: {', '.join(missing_match_fields)}")

    rel_ws = wb["18_关联关系"]
    workbook_relation_codes = {str(rel_ws.cell(row=row, column=1).value or "").strip() for row in range(2, rel_ws.max_row + 1)}
    missing_workbook_relations = sorted(EXPECTED_RELATIONS - workbook_relation_codes)
    if missing_workbook_relations:
        fail(f"18_关联关系 missing rows: {', '.join(missing_workbook_relations)}")

    if not GRAPH_JSON.exists():
        fail(f"missing graph json: {GRAPH_JSON}")
    if not GRAPH_HTML.exists():
        fail(f"missing graph html: {GRAPH_HTML}")
    data = json.loads(GRAPH_JSON.read_text(encoding="utf-8"))
    object_keys = {item["key"] for item in data.get("objects", [])}
    if "industry_node_relation" not in object_keys:
        fail("dw_schema.json missing industry_node_relation object")
    if "major_industry_node_match" not in object_keys:
        fail("dw_schema.json missing major_industry_node_match object")

    graph_relation_codes = {item["code"] for item in data.get("relations", [])}
    missing_graph_relations = sorted(EXPECTED_RELATIONS - graph_relation_codes)
    if missing_graph_relations:
        fail(f"dw_schema.json missing relations: {', '.join(missing_graph_relations)}")

    html = GRAPH_HTML.read_text(encoding="utf-8")
    if 'data-source="major" data-target="industry_node"' not in html:
        fail("dw_er.html missing visible direct edge from major to industry_node")
    viewbox_match = re.search(r'<svg class="graph" viewBox="0 0 (\d+) (\d+)"', html)
    if not viewbox_match:
        fail("dw_er.html missing graph viewBox")
    width, height = (int(viewbox_match.group(1)), int(viewbox_match.group(2)))
    if width < 2020 or height < 1230:
        fail(f"dw_er.html viewBox too small: {width}x{height}")

    print("PASS: industry node relation DW model is present")


if __name__ == "__main__":
    main()
