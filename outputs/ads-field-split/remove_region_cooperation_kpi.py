from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries


ROOT = Path("/Users/liuhongzhe/Desktop/学堂/专业建设/Codex工程")
WORKBOOK = ROOT / "V1.0需求（2026.6.11）/官方数据/【ADS】产业布局-产业链图谱数据源梳理.xlsx"
SHEET_NAME = "ADS表"
TARGET_SECONDARY = "区域KPI卡片"
REMOVED_FIELDS = {"7. 合作线索数量", "8. 合作线索类型说明"}


def merged_block(ws, row: int, col: int) -> tuple[int, int]:
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_col <= col <= merged_range.max_col
            and merged_range.min_row <= row <= merged_range.max_row
        ):
            return merged_range.min_row, merged_range.max_row
    return row, row


def shift_range(coord: str, delete_at: int, amount: int) -> str | None:
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    delete_end = delete_at + amount - 1
    if max_row < delete_at:
        shifted = (min_col, min_row, max_col, max_row)
    elif min_row > delete_end:
        shifted = (min_col, min_row - amount, max_col, max_row - amount)
    elif min_row < delete_at <= max_row:
        shifted = (min_col, min_row, max_col, max_row - amount)
    else:
        return None

    shifted_min_col, shifted_min_row, shifted_max_col, shifted_max_row = shifted
    if shifted_min_row > shifted_max_row:
        return None
    return (
        f"{get_column_letter(shifted_min_col)}{shifted_min_row}:"
        f"{get_column_letter(shifted_max_col)}{shifted_max_row}"
    )


def shift_image_anchor(image, delete_at: int, amount: int) -> None:
    anchor = image.anchor
    delete_index = delete_at - 1
    for marker_name in ("_from", "to"):
        marker = getattr(anchor, marker_name, None)
        if marker is not None and getattr(marker, "row", 0) >= delete_index:
            marker.row -= amount


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb[SHEET_NAME]

    target_start = target_end = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, 3).value or "").strip() == TARGET_SECONDARY:
            target_start, target_end = merged_block(ws, row, 3)
            break

    if target_start is None or target_end is None:
        raise RuntimeError(f"未找到 {TARGET_SECONDARY}")

    current_fields = [str(ws.cell(row, 4).value or "").strip() for row in range(target_start, target_end + 1)]
    if set(current_fields[-2:]) != REMOVED_FIELDS:
        raise RuntimeError(f"{TARGET_SECONDARY} 末尾字段不是预期的合作线索字段: {current_fields[-2:]}")

    delete_at = target_end - 1
    amount = 2
    merged_ranges = [str(merged_range) for merged_range in ws.merged_cells.ranges]
    shifted_ranges = [
        shifted_range
        for coord in merged_ranges
        if (shifted_range := shift_range(coord, delete_at, amount)) is not None
    ]

    for coord in merged_ranges:
        ws.unmerge_cells(coord)

    ws.delete_rows(delete_at, amount)

    for image in ws._images:
        shift_image_anchor(image, delete_at, amount)

    for coord in shifted_ranges:
        ws.merge_cells(coord)

    wb.save(WORKBOOK)
    print(f"removed {amount} rows from {TARGET_SECONDARY}: {delete_at}-{delete_at + amount - 1}")
    print(f"rows={ws.max_row} images={len(ws._images)} merged={len(ws.merged_cells.ranges)}")


if __name__ == "__main__":
    main()
