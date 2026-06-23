from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell


ROOT = Path("/Users/liuhongzhe/Desktop/学堂/专业建设/Codex工程")
WORKBOOK = ROOT / "V1.0需求（2026.6.11）/官方数据/【ADS】产业布局-产业链图谱数据源梳理.xlsx"


DATABASE_FIELDS: dict[str, list[str]] = {
    "页面顶部研判摘要": ["页面名称", "产业链名称", "页面目标说明", "研判摘要正文"],
    "国标行业KPI卡片": ["指标名称", "指标数值", "指标单位", "指标说明"],
    "国标行业KPI详情弹窗": [
        "指标名称",
        "国标行业分类提示",
        "指标说明",
        "指标数值",
        "详情摘要",
        "统计口径",
        "关键指标项",
        "关联行业名称",
        "专业建设提示",
        "指标更新时间",
    ],
    "矩形树图-阶段分布": ["阶段编码", "阶段名称", "阶段说明", "阶段统计值", "覆盖国标行业"],
    "矩形树图-产业节点": ["产业节点编码", "所属阶段", "产业节点名称", "代表企业数量", "产品技术服务标签", "展示权重值"],
    "桑基图-流向与hover详情": [
        "产业节点列序号",
        "产业节点名称",
        "节点企业数量",
        "技术领域标签",
        "来源产业环节编码",
        "目标产业环节编码",
        "关系强度值",
        "关系说明",
        "详情摘要",
    ],
    "国标行业关联分析-代表企业行业覆盖": ["行业门类", "细分行业名称", "覆盖占比", "代表企业数量", "样本企业名称", "覆盖指数值"],
    "国标行业关联分析-行业增长信号": ["行业名称", "招聘热度等级", "政策热度等级", "企业活跃度等级", "增长信号说明"],
    "产业链洞察卡片": ["洞察标题", "洞察正文", "适用产业环节", "专业建设动作"],
    "代表企业列表": ["产业阶段", "企业名称", "企业对接能力", "产品技术服务节点"],
    "核心岗位列表": ["产业阶段", "岗位名称", "岗位所属环节", "岗位入口标识"],
    "产业链建设建议": ["建议序号", "建议标题", "建议描述"],
    "区域KPI卡片": ["覆盖省份数量", "企业样本数量", "重点城市数量"],
    "区域合作方向卡片": ["区域名称", "产业方向", "合作资源说明", "实训方向描述"],
    "政策筛选工具栏": ["政策级别", "政策关键词", "政策来源类型", "政策发布日期范围"],
    "政策时间线列表": [
        "政策日期",
        "发布日期",
        "政策标题",
        "政策级别",
        "政策标签",
        "发布机构",
        "政策来源",
        "原文链接",
        "政策描述",
        "政策摘要",
        "影响分析",
        "专业建设任务",
    ],
    "政策关键词云": ["政策关键词", "关键词热度值", "关键词分类"],
    "年度政策趋势": ["统计年份", "政策关注度值"],
    "企业搜索交互": ["企业检索关键词", "检索范围", "匹配企业数量"],
    "企业列表表格": ["企业名称", "统一社会信用代码", "注册地址", "企业规模", "产品技术服务节点", "所属产业"],
    "分页与空状态": ["当前页码", "每页记录数", "总页数", "总记录数", "无结果提示文案"],
    "岗位分析分组与页面切换": ["功能分组名称", "分析页面名称", "当前分析模式", "当前分析标签"],
    "KPI与岗位搜索": ["指标名称", "指标数值", "指标单位", "指标色调", "岗位检索关键词"],
    "岗位列表、等级筛选与分页": [
        "岗位等级筛选值",
        "岗位等级选项",
        "过滤后岗位数量",
        "岗位名称",
        "岗位等级",
        "薪资区间",
        "需求等级",
        "产业链环节",
        "技能关键词",
        "当前页码",
        "每页记录数",
        "总记录数",
    ],
    "岗位画像详情弹窗": [
        "岗位名称",
        "数据来源说明",
        "薪资区间",
        "薪资单位",
        "学历要求",
        "经验要求",
        "岗位层级",
        "职业发展路径",
        "岗位概述",
        "岗位标签",
        "三维能力摘要",
        "典型工作任务",
        "推荐职业资格证书",
        "对接专业",
        "相关企业",
        "详情更新时间",
        "能力图谱关联标识",
    ],
    "三维能力分析与雷达图": ["能力组名称", "能力项列表", "能力数量", "能力维度标签", "能力得分序列", "能力分类颜色"],
    "证书入口与详情数据缺口": ["证书编码", "证书名称", "证书等级", "颁发机构", "考试科目", "适用岗位", "适配专业", "详情配置状态"],
    "相关企业详情弹窗": [
        "企业名称",
        "企业全称",
        "企业标签",
        "企业摘要",
        "所在地区",
        "所属行业",
        "企业规模",
        "成立时间",
        "年均招聘量",
        "产业链环节",
        "核心产品",
        "技术方向",
        "主要招聘岗位",
        "校企合作类型",
        "合作详情",
    ],
    "岗位能力图谱入口": ["岗位编码", "任务节点名称", "能力节点名称", "任务能力关系", "激活任务名称", "能力侧栏摘要"],
    "需求KPI卡片": ["指标名称", "指标数值", "指标单位", "趋势说明"],
    "岗位需求月度趋势": ["统计月份", "需求指数值"],
    "技能需求热度": ["技能名称", "热度百分比"],
    "热门岗位招聘明细表": ["岗位名称", "招聘需求量", "薪资区间", "增长趋势", "重点城市"],
    "新兴技术方向卡片": ["技术方向名称", "成熟阶段", "产业影响", "关联岗位", "推荐专业"],
    "新岗位×专业匹配": ["岗位名称", "紧缺度", "薪资区间", "职业路径", "匹配专业", "相关专业", "推荐能力"],
    "人才培养方向建议表": ["技术方向", "建议课程", "相关专业"],
}


def get_merged_block_start(ws, row: int, col: int) -> int:
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col <= col <= merged_range.max_col and merged_range.min_row <= row <= merged_range.max_row:
            return merged_range.min_row
    return row


def current_value(ws, row: int, col: int) -> str:
    start = get_merged_block_start(ws, row, col)
    value = ws.cell(start, col).value
    return "" if value is None else str(value).strip()


def clear_field_block(ws, start: int, end: int) -> None:
    for row in range(start, end + 1):
        cell = ws.cell(row, 4)
        if not isinstance(cell, MergedCell):
            cell.value = None


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["ADS表"]

    groups: list[tuple[int, int, str]] = []
    row = 2
    while row <= ws.max_row:
        start = get_merged_block_start(ws, row, 3)
        end = row
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= 3 <= merged_range.max_col and merged_range.min_row == start:
                end = merged_range.max_row
                break
        secondary = current_value(ws, start, 3)
        groups.append((start, end, secondary))
        row = end + 1

    updated = 0
    for start, end, secondary in groups:
        fields = DATABASE_FIELDS.get(secondary)
        if not fields:
            continue
        clear_field_block(ws, start, end)
        for offset, field_name in enumerate(fields):
            target_row = start + offset
            if target_row > end:
                break
            ws.cell(target_row, 4).value = f"{offset + 1}. {field_name}"
        updated += 1

    wb.save(WORKBOOK)
    print(f"updated_groups={updated}")


if __name__ == "__main__":
    main()
