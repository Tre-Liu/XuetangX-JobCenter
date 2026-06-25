from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell


ROOT = Path("/Users/liuhongzhe/Desktop/学堂/专业建设/Codex工程")
WORKBOOK = ROOT / "V1.0需求（2026.6.11）/官方数据/【ADS】产业布局-产业链图谱数据源梳理.xlsx"


FIELD_REPLACEMENTS_BY_SECONDARY = {
    "矩形树图-阶段分布": ["阶段编码", "阶段名称", "产业类数量", "代表企业数量", "阶段说明", "覆盖国标行业", "阶段排序"],
    "国标行业关联分析-代表企业行业覆盖": ["行业门类名称", "细分行业名称", "行业覆盖占比", "代表企业数量", "样本企业名称", "覆盖说明"],
    "核心岗位列表": ["产业阶段", "岗位名称", "岗位所属环节", "岗位需求等级", "薪资区间", "重点城市", "相关能力标签", "岗位关联标识"],
    "区域KPI卡片": ["覆盖省份数量", "覆盖区域范围说明", "产业链相关企业样本数量", "企业样本统计口径", "重点城市数量", "产业集聚城市名称"],
    "技能需求热度": ["技能名称", "技能热度百分比", "技能类别", "关联岗位方向", "关联产业环节", "数据统计周期"],
    "新兴技术方向卡片": ["技术方向名称", "成熟阶段", "产业影响", "关联岗位", "推荐专业", "能力要求", "课程实训建议", "数据来源说明"],
}


def merged_block(ws, row: int, col: int) -> tuple[int, int]:
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col <= col <= merged_range.max_col and merged_range.min_row <= row <= merged_range.max_row:
            return merged_range.min_row, merged_range.max_row
    return row, row


def set_fields(ws, start: int, end: int, fields: list[str]) -> None:
    for index, row in enumerate(range(start, end + 1), start=1):
        cell = ws.cell(row, 4)
        if isinstance(cell, MergedCell):
            continue
        if index <= len(fields):
            cell.value = f"{index}. {fields[index - 1]}"
        else:
            cell.value = None


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["ADS表"]
    updated = []
    row = 2
    while row <= ws.max_row:
        start, end = merged_block(ws, row, 3)
        if row != start:
            row = end + 1
            continue
        secondary = str(ws.cell(start, 3).value or "").strip()
        fields = FIELD_REPLACEMENTS_BY_SECONDARY.get(secondary)
        if fields:
            set_fields(ws, start, end, fields)
            updated.append((secondary, start, end))
        row = end + 1

    wb.save(WORKBOOK)
    for secondary, start, end in updated:
        print(f"updated {secondary}: {start}-{end}")


if __name__ == "__main__":
    main()
