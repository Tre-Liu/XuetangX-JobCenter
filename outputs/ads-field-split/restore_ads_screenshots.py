from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image


ROOT = Path("/Users/liuhongzhe/Desktop/学堂/专业建设/Codex工程")
BACKUP = ROOT / "V1.0需求（2026.6.11）/官方数据/【ADS】产业布局-产业链图谱数据源梳理_原始备份_20260623.xlsx"
CURRENT = ROOT / "V1.0需求（2026.6.11）/官方数据/【ADS】产业布局-产业链图谱数据源梳理.xlsx"


def clean_field_item(value: str) -> str:
    text = str(value or "")
    for prefix in ["字段包括", "字段按", "字段", "弹窗字段包括", "交互包括", "当前展示", "当前包含", "当前包括", "展示", "和", "及", "以及"]:
        if text.startswith(prefix):
            text = text[len(prefix) :]
    return text.strip("。；;，,、 \n\t")


def split_list_text(text: str) -> list[str]:
    import re

    return [clean_field_item(item) for item in re.split(r"[、,，]", text) if clean_field_item(item) and clean_field_item(item) != "等"]


def first_sentence_after(text: str, marker: str) -> str:
    idx = text.find(marker)
    if idx < 0:
        return ""
    return text[idx + len(marker) :].split("。")[0]


def trim_trailing_explanation(text: str) -> str:
    for marker in ["，并", "，用于", "，当前", "，如", "，例如", "，点击"]:
        text = text.split(marker)[0]
    return text.strip()


def extract_fields(description: str) -> list[str]:
    import re

    text = re.sub(r"\s+", " ", str(description or "")).strip()
    if not text:
        return [""]
    for marker in ["弹窗字段包括", "字段包括", "交互包括", "字段为"]:
        fields = split_list_text(trim_trailing_explanation(first_sentence_after(text, marker)))
        if len(fields) > 1:
            return fields
    match = re.search(r"字段按(.+?)分组[，,](?:展示)?(.+?)(?:。|，如|，例如|$)", text)
    if match:
        fields = split_list_text(match.group(1)) + split_list_text(match.group(2))
        if len(fields) > 1:
            return fields
    match = re.search(r"(?:包含|可展示)(.+?)(?:。|，|$)", text)
    if match:
        fields = split_list_text(match.group(1))
        if 1 < len(fields) <= 12:
            return fields
    return [text]


def build_row_map(backup_ws) -> dict[int, tuple[int, int]]:
    mapping: dict[int, tuple[int, int]] = {}
    current_row = 2
    for source_row in range(2, backup_ws.max_row + 1):
        fields = extract_fields(backup_ws.cell(source_row, 4).value)
        start = current_row
        current_row += len(fields)
        mapping[source_row] = (start, current_row - 1)
    return mapping


def main() -> None:
    backup_wb = openpyxl.load_workbook(BACKUP)
    backup_ws = backup_wb["ADS表"]
    current_wb = openpyxl.load_workbook(CURRENT)
    current_ws = current_wb["ADS表"]

    images_by_row: dict[int, list] = defaultdict(list)
    for image in backup_ws._images:
        row = image.anchor._from.row + 1
        images_by_row[row].append(image)

    # Remove any images introduced by previous repair runs and clear the text placeholders.
    current_ws._images = []
    for row in range(2, current_ws.max_row + 1):
        cell = current_ws.cell(row, 5)
        if not isinstance(cell, MergedCell):
            cell.value = None

    row_map = build_row_map(backup_ws)
    restored = 0
    for source_row, images in images_by_row.items():
        if source_row not in row_map:
            continue
        image = max(images, key=lambda item: item.width * item.height)
        start_row, end_row = row_map[source_row]
        height_rows = max(1, end_row - start_row + 1)

        restored_image = Image(BytesIO(image._data()))
        max_width = 116
        max_height = min(180, max(42, height_rows * 23))
        scale = min(max_width / restored_image.width, max_height / restored_image.height, 1)
        restored_image.width = int(restored_image.width * scale)
        restored_image.height = int(restored_image.height * scale)
        restored_image.anchor = f"E{start_row}"
        current_ws.add_image(restored_image)

        current_ws.row_dimensions[start_row].height = max(current_ws.row_dimensions[start_row].height or 18, min(120, restored_image.height * 0.75 + 6))
        restored += 1

    current_wb.save(CURRENT)
    print(f"restored_images={restored}")


if __name__ == "__main__":
    main()
