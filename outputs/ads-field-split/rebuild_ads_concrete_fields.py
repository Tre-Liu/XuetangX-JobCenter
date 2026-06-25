from __future__ import annotations

from collections import defaultdict
from copy import copy
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path("/Users/liuhongzhe/Desktop/学堂/专业建设/Codex工程")
BACKUP = ROOT / "V1.0需求（2026.6.11）/官方数据/【ADS】产业布局-产业链图谱数据源梳理_原始备份_20260623.xlsx"
CURRENT = ROOT / "V1.0需求（2026.6.11）/官方数据/【ADS】产业布局-产业链图谱数据源梳理.xlsx"


FIELDS_BY_SOURCE_ROW: dict[int, list[str]] = {
    2: ["页面名称", "产业链名称", "当前产业链编码", "研判标题", "上游能力摘要", "中游服务摘要", "下游应用摘要", "专业建设组织建议"],
    3: ["关联国标行业数量", "关联国标行业覆盖层级", "覆盖门类数量", "覆盖门类名称", "核心关联行业名称", "岗位企业样本集中说明", "增长行业名称", "增长信号说明"],
    4: ["指标类型", "国标行业分类版本", "关联行业门类", "关联行业代码", "指标值", "样本企业数量", "统计口径说明", "专业建设提示"],
    5: ["上游产业类数", "上游代表企业数", "中游产业类数", "中游代表企业数", "下游产业类数", "下游代表企业数", "覆盖国标行业"],
    6: ["产业节点编码", "所属阶段", "产业节点名称", "代表企业数量", "产品技术服务标签", "节点展示权重"],
    7: ["来源产业环节编码", "来源产业环节名称", "目标产业环节编码", "目标产业环节名称", "关系强度值", "关系类型", "关系说明", "节点企业数量", "技术领域标签"],
    8: ["建筑业覆盖占比", "建筑业代表企业数", "软件和信息技术服务业覆盖占比", "制造业覆盖占比", "科学研究和技术服务业覆盖占比", "样本企业名称"],
    9: ["行业名称", "招聘热度等级", "政策热度等级", "企业活跃度等级", "增长原因说明", "专业建设启示"],
    10: ["价值流判断", "建设切入点", "企业反馈"],
    11: ["产业阶段", "企业名称", "企业对接能力", "产品技术服务节点"],
    12: ["上游岗位名称", "中游岗位名称", "下游岗位名称", "智能测量工程师", "BIM深化设计工程师", "智慧建造平台实施顾问", "智能建造施工技术员", "结构健康监测工程师"],
    13: ["建议序号", "建议标题", "建议描述"],
    14: ["页面名称", "当前产业链名称", "区域布局研判摘要", "区域岗位需求说明"],
    15: ["覆盖省份数量", "覆盖区域范围说明", "智能建造相关企业样本数量", "企业样本统计口径", "重点城市数量", "产业集聚城市名称"],
    16: ["区域名称", "产业方向", "企业工程场景", "校企合作方向", "实训基地任务"],
    17: ["页面名称", "当前产业链名称", "政策趋势摘要", "政策信号方向"],
    18: ["政策级别", "政策关键词", "政策发布时间范围"],
    19: ["政策发布日期", "政策标题", "政策级别", "政策标签", "发布机构", "政策来源", "原文链接", "政策描述", "政策摘要", "影响分析", "专业建设任务", "关联产业方向"],
    20: ["政策关键词", "关键词热度值", "关键词分类"],
    21: ["统计年份", "政策关注度值"],
    22: ["页面名称", "当前产业链名称", "企业资源研判摘要", "企业资源用途说明"],
    23: ["企业检索关键词", "检索字段范围", "匹配企业数"],
    24: ["企业名称", "统一社会信用代码", "注册地址", "企业规模", "产品技术服务节点", "所属产业"],
    25: ["当前页码", "每页记录数", "总页数", "总记录数", "无结果提示文案"],
    26: ["分析模块编码", "分析模块名称", "一级功能名称", "二级功能名称"],
    27: ["页面名称", "当前产业链名称", "岗位画像洞察摘要", "课程体系补强建议"],
    28: ["岗位数量", "岗位数量变化趋势", "典型工作任务数量", "典型工作任务单位", "能力项数量", "能力项统计口径", "证书数量", "证书资源类型", "岗位检索关键词", "产业链环节筛选值"],
    29: ["岗位等级筛选值", "岗位等级选项", "过滤后岗位数量", "岗位名称", "岗位等级", "薪资区间", "需求等级", "产业链环节", "技能关键词", "当前页码", "总页数"],
    30: ["岗位名称", "数据来源说明", "薪资区间", "薪资单位", "学历要求", "经验要求", "岗位层级", "职业发展路径", "岗位概述", "岗位标签", "知识能力摘要", "技能能力摘要", "素养能力摘要", "典型工作任务", "推荐职业资格证书", "对接专业", "相关企业"],
    31: ["知识能力项", "技能能力项", "素养能力项", "能力项数量", "雷达图维度名称", "能力得分"],
    32: ["证书编码", "证书名称", "证书等级", "颁发机构", "考试科目", "适用岗位", "适配专业", "证书详情配置状态"],
    33: ["企业名称", "企业全称", "企业标签", "企业摘要", "所在地区", "所属行业", "企业规模", "成立时间", "年均招聘量", "产业链环节", "核心产品", "技术方向", "主要招聘岗位", "校企合作类型", "合作详情"],
    34: ["岗位编码", "典型任务名称", "知识能力项", "技能能力项", "素养能力项", "任务能力关系", "课程训练建议"],
    35: ["页面名称", "当前产业链名称", "招聘需求趋势摘要", "岗位能力包建设建议"],
    36: ["近十二月岗位需求量", "岗位需求趋势值", "高频招聘岗位数量", "高频岗位统计口径", "平均薪资金额", "薪资单位", "薪资变化趋势值", "企业样本数量", "企业样本来源说明", "企业样本趋势值"],
    37: ["统计月份", "岗位需求量", "需求指数", "同比变化值"],
    38: ["BIM建模与深化设计热度", "智慧工地平台应用热度", "智能测量与三维扫描热度", "智能检测监测数据分析热度", "装配式构件深化与生产热度", "建筑机器人与设备联调热度"],
    39: ["岗位名称", "招聘需求量", "薪资区间", "增长趋势", "重点城市"],
    40: ["页面名称", "当前产业链名称", "新岗位新技术摘要", "课程实训布局建议"],
    41: ["建筑机器人协同施工", "BIM数字孪生工地", "结构健康智能监测", "装配式构件数字工厂", "成熟阶段", "产业影响", "关联岗位", "推荐专业"],
    42: ["岗位名称", "紧缺度", "薪资区间", "职业路径", "匹配专业", "相关专业", "推荐能力"],
    43: ["技术方向", "建议课程", "相关专业"],
}


BLOCKS_BY_SOURCE_ROW: dict[int, list[tuple[str, list[str]]]] = {
    4: [
        (
            "关联国标行业详情弹窗",
            [
                "关联行业数量",
                "行业分类层级范围",
                "关联行业门类名称",
                "关联行业大类名称",
                "关联行业中类名称",
                "国标行业分类版本",
                "统计样本口径",
                "专业建设使用说明",
            ],
        ),
        (
            "覆盖门类详情弹窗",
            [
                "覆盖门类数量",
                "覆盖门类名称",
                "覆盖门类代码",
                "主要门类企业样本数",
                "门类覆盖说明",
                "关联岗位方向",
                "专业建设使用说明",
            ],
        ),
        (
            "核心关联行业详情弹窗",
            [
                "核心关联行业名称",
                "核心关联行业代码",
                "岗位企业样本集中说明",
                "代表企业数量",
                "关联产业节点",
                "关联岗位方向",
                "专业建设使用说明",
            ],
        ),
        (
            "增长行业详情弹窗",
            [
                "增长行业名称",
                "增长行业代码",
                "招聘热度等级",
                "政策热度等级",
                "企业活跃度等级",
                "增长信号说明",
                "专业建设启示",
            ],
        ),
    ]
}


HEADERS = ["模块", "一级功能", "二级功能", "字段描述", "截图", "是否AI输出", "建议提示词"]

SECONDARY_NAME_BY_SOURCE_ROW = {
    7: "桑基图-产业环节关系详情",
    18: "政策筛选条件",
    23: "企业检索条件",
    25: "分页状态数据",
    26: "岗位分析页面配置",
    32: "证书详情数据缺口",
    34: "岗位能力图谱数据",
}


PROMPT_BY_SOURCE_ROW = {
    30: "请基于【岗位名称】【薪资区间】【需求量】【产业链节点】【能力项】【典型任务】【推荐证书】【对接专业】【相关企业】生成岗位画像详情文案。要求输出岗位概述、三维能力、典型任务、证书建议、专业对接和企业支撑，不夸大招聘数据；证书、专业、企业匹配需能回溯到岗位资源匹配库。",
    32: "请基于【证书名称】【等级】【颁发机构】【考试科目】【适用岗位】【适配专业】【岗位能力要求】生成证书详情数据。要求补齐证书详情结构，并说明证书与岗位能力、课程模块、实训评价之间的关系，不夸大证书效力。",
    33: "请基于【企业名称】【所属产业】【产品/技术/服务节点】【代表岗位】【岗位能力】【合作场景】生成企业资源研判摘要。要求说明企业能为岗位画像、课程案例、实训项目和校企合作提供什么支撑，并保留数据来源边界。",
    40: "请基于【新技术方向】【成熟度/阶段】【产业影响】【新增岗位】【职业路径】【相关专业】【推荐能力】生成新岗位新技术研判。要求输出技术趋势、人才缺口、职业发展路径、课程实训建议和专业协同方向。",
    41: "请基于【新技术方向】【成熟度/阶段】【产业影响】【新增岗位】【职业路径】【相关专业】【推荐能力】生成新岗位新技术研判。要求输出技术趋势、人才缺口、职业发展路径、课程实训建议和专业协同方向。",
    42: "请基于【新技术方向】【成熟度/阶段】【产业影响】【新增岗位】【职业路径】【相关专业】【推荐能力】生成新岗位新技术研判。要求输出技术趋势、人才缺口、职业发展路径、课程实训建议和专业协同方向。",
    43: "请基于【新技术方向】【成熟度/阶段】【产业影响】【新增岗位】【职业路径】【相关专业】【推荐能力】生成新岗位新技术研判。要求输出技术趋势、人才缺口、职业发展路径、课程实训建议和专业协同方向。",
}


DW_SOURCE_ROW = [
    "DW字段模型",
    "DW_专业建设数据模型设计_岗位所属行业更新.xlsx",
    "V1.0需求（2026.6.11）/官方数据/DW_专业建设数据模型设计_岗位所属行业更新.xlsx",
    "作为岗位画像推荐证书、对接专业、相关企业，以及新技术方向、职业路径、课程、能力项的字段支撑说明",
    "岗位资源匹配和新技术研判类字段仍需AI生成后审核和专家复核",
]


def style_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    widths = {"A": 14, "B": 16, "C": 26, "D": 36, "E": 15, "F": 12, "G": 70}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    header_fill = PatternFill("solid", fgColor="115F7A")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="000000")
    side = Side(style="thin", color="D7DEE8")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24


def ensure_source_index(wb) -> None:
    if "来源索引" not in wb.sheetnames:
        return
    ws = wb["来源索引"]
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, 2).value or "").strip() == DW_SOURCE_ROW[1]:
            for col, value in enumerate(DW_SOURCE_ROW, 1):
                ws.cell(row, col).value = value
            return
    target_row = ws.max_row + 1
    for col, value in enumerate(DW_SOURCE_ROW, 1):
        src = ws.cell(max(2, target_row - 1), col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.value = value
    ws.row_dimensions[target_row].height = ws.row_dimensions[max(2, target_row - 1)].height


def image_by_source_row(ws) -> dict[int, tuple[bytes, int, int]]:
    images: dict[int, list] = defaultdict(list)
    for image in ws._images:
        images[image.anchor._from.row + 1].append(image)
    result: dict[int, tuple[bytes, int, int]] = {}
    for row, items in images.items():
        image = max(items, key=lambda item: item.width * item.height)
        result[row] = (image._data(), image.width, image.height)
    return result


def rebuild_ads_sheet() -> None:
    backup_wb = openpyxl.load_workbook(BACKUP)
    backup_ws = backup_wb["ADS表"]
    images = image_by_source_row(backup_ws)

    wb = openpyxl.load_workbook(CURRENT)
    if "ADS表" in wb.sheetnames:
        del wb["ADS表"]
    ws = wb.create_sheet("ADS表", 0)
    ws.append(HEADERS)
    ensure_source_index(wb)

    row_cursor = 2
    merge_blocks: list[tuple[int, int, int]] = []
    for source_row in range(2, backup_ws.max_row + 1):
        default_secondary = SECONDARY_NAME_BY_SOURCE_ROW.get(source_row, backup_ws.cell(source_row, 3).value)
        blocks = BLOCKS_BY_SOURCE_ROW.get(source_row) or [(default_secondary, FIELDS_BY_SOURCE_ROW[source_row])]
        for secondary_name, fields in blocks:
            start = row_cursor
            for idx, field in enumerate(fields, 1):
                ws.append(
                    [
                        backup_ws.cell(source_row, 1).value,
                        backup_ws.cell(source_row, 2).value,
                        secondary_name,
                        f"{idx}. {field}",
                        None,
                        backup_ws.cell(source_row, 6).value,
                        PROMPT_BY_SOURCE_ROW.get(source_row, backup_ws.cell(source_row, 7).value),
                    ]
                )
                row_cursor += 1
            end = row_cursor - 1
            merge_blocks.append((source_row, start, end))

    style_sheet(ws)
    for source_row, start, end in merge_blocks:
        if end > start:
            for col in [1, 2, 3, 5, 6, 7]:
                ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)
        for row in range(start, end + 1):
            ws.row_dimensions[row].height = 20

        image_data = images.get(source_row)
        if image_data:
            data, original_width, original_height = image_data
            restored = Image(BytesIO(data))
            max_width = 112
            max_height = min(180, max(42, (end - start + 1) * 22))
            scale = min(max_width / original_width, max_height / original_height, 1)
            restored.width = int(original_width * scale)
            restored.height = int(original_height * scale)
            restored.anchor = f"E{start}"
            ws.add_image(restored)
            ws.row_dimensions[start].height = max(ws.row_dimensions[start].height or 20, min(120, restored.height * 0.75 + 6))

    wb.save(CURRENT)
    print(f"rows={ws.max_row} images={len(ws._images)}")


if __name__ == "__main__":
    rebuild_ads_sheet()
